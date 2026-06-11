from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import shutil
from extractor import extract_text
from llm import extract_terms_with_ollama, extract_terms_with_cancellation, cancel_processing, reset_cancel
from fastapi.responses import FileResponse, Response
import asyncio
from concurrent.futures import ThreadPoolExecutor

executor = ThreadPoolExecutor(max_workers=1)


load_dotenv()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

try:
    conn = psycopg2.connect(
        host=os.getenv("DB_HOST"),
        database=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=os.getenv("DB_PORT")
    )
    conn.autocommit = True
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    print("Database connected successfully")
except Exception as e:
    print(f"Database connection failed: {e}")

UPLOAD_DIR = "uploads"
STORE_DIR = "stored"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(STORE_DIR, exist_ok=True)

def normalize(text: str) -> str:
    text = text.lower().strip()
    text = text.replace("the laws of", "laws of")
    text = text.replace("the courts of", "courts of")
    text = " ".join(text.split())
    return text

def normalize_term_key(term: str) -> str:
    term = term.lower().strip()
    term = re.sub(r'\s*\(.*?\)', '', term)
    term = " ".join(term.split())
    return term

def is_valid_term(term: dict) -> bool:
    invalid_values = {
        "", "null", "none", "n/a", "—", "-",
        "unknown", "not stated", "not specified",
        "not applicable", "not explicitly stated",
        "not provided", "not mentioned"
    }
    if "term" not in term or "value" not in term:
        return False
    if term["term"] is None or term["value"] is None:
        return False
    if str(term["term"]).strip() == "":
        return False
    if str(term["value"]).strip().lower() in invalid_values:
        return False
    return True

@app.get("/")
def root():
    return {"message": "Doc Processor API is running"}


@app.post("/api/documents/upload")
async def upload_document(request: Request, file: UploadFile = File(...)):
    reset_cancel()
    filename = file.filename
    ext = filename.split(".")[-1].lower()

    if ext not in ["pdf", "docx", "txt"]:
        raise HTTPException(
            status_code=400,
            detail="Only PDF, DOCX and TXT files are supported"
        )

    upload_path = os.path.join(UPLOAD_DIR, filename)
    store_path = os.path.join(STORE_DIR, filename)

    with open(upload_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    shutil.copy2(upload_path, store_path)

    try:
        if await request.is_disconnected():
            return {"message": "Cancelled"}

        print(f"Extracting text from {filename}...")
        text = extract_text(upload_path, ext)
        print(f"Extracted {len(text)} characters")

        if await request.is_disconnected():
            return {"message": "Cancelled"}

        print("Sending to Ollama for term extraction...")

        # run blocking Ollama calls in thread pool
        # this frees FastAPI event loop to handle cancel requests
        loop = asyncio.get_event_loop()
        terms = await loop.run_in_executor(
            executor,
            lambda: extract_terms_with_cancellation(text)
        )

        if terms is None:
            print("Processing cancelled")
            if os.path.exists(store_path):
                os.remove(store_path)
            return {"message": "Cancelled"}

        print(f"Extracted {len(terms)} terms")

        if not terms:
            raise HTTPException(
                status_code=500,
                detail="Failed to extract terms from document"
            )

        cursor.execute(
            "INSERT INTO documents (filename, file_type, file_path) VALUES (%s, %s, %s) RETURNING *",
            (filename, ext, store_path)
        )
        document = cursor.fetchone()
        document_id = document["id"]

        for term in terms:
            if is_valid_term(term):
                cursor.execute(
                    "INSERT INTO terms (document_id, term, value) VALUES (%s, %s, %s)",
                    (document_id, term["term"], term["value"])
                )

        

        return {
            "document": document,
            "terms": terms,
            "total": len(terms)
        }

    except HTTPException:
        raise
    except Exception as e:
        
        if os.path.exists(store_path):
            os.remove(store_path)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if os.path.exists(upload_path):
            os.remove(upload_path)

@app.get("/api/documents")
def get_documents():
    try:
        cursor.execute(
            "SELECT * FROM documents ORDER BY uploaded_at DESC"
        )
        documents = cursor.fetchall()
        return documents
    except Exception as e:
        
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents/{document_id}/terms")
def get_terms(document_id: int):
    try:
        cursor.execute(
            "SELECT * FROM documents WHERE id = %s",
            (document_id,)
        )
        document = cursor.fetchone()

        if not document:
            raise HTTPException(
                status_code=404,
                detail="Document not found"
            )

        cursor.execute(
            "SELECT term, value FROM terms WHERE document_id = %s ORDER BY id",
            (document_id,)
        )
        terms = cursor.fetchall()

        return {
            "document": document,
            "terms": terms,
            "total": len(terms)
        }
    except HTTPException:
        raise
    except Exception as e:
        
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/documents/{document_id}")
def delete_document(document_id: int):
    try:
        cursor.execute(
            "SELECT * FROM documents WHERE id = %s",
            (document_id,)
        )
        existing = cursor.fetchone()

        if not existing:
            raise HTTPException(
                status_code=404,
                detail="Document not found"
            )

        file_path = existing.get("file_path")
        if file_path and os.path.exists(file_path):
            os.remove(file_path)

        cursor.execute(
            "DELETE FROM documents WHERE id = %s",
            (document_id,)
        )

        return {"message": "Document deleted"}
    except HTTPException:
        raise
    except Exception as e:
        
        raise HTTPException(status_code=500, detail=str(e))
    


@app.get("/api/documents/{document_id}/file")
def get_document_file(document_id: int):
    try:
        cursor.execute(
            "SELECT * FROM documents WHERE id = %s",
            (document_id,)
        )
        document = cursor.fetchone()

        if not document:
            raise HTTPException(
                status_code=404,
                detail="Document not found"
            )

        file_path = document.get("file_path")

        if not file_path or not os.path.exists(file_path):
            raise HTTPException(
                status_code=404,
                detail="Source file not available — please re-upload the document"
            )

        ext = document["file_type"]

        media_types = {
            "pdf":  "application/pdf",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "txt":  "text/plain"
        }

        # inline tells browser to display, not download
        return FileResponse(
            path=file_path,
            filename=document["filename"],
            media_type=media_types.get(ext, "application/octet-stream"),
            headers={
                "Content-Disposition": f"inline; filename=\"{document['filename']}\""
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents/{document_id}/export/excel")
def export_excel(document_id: int):
    try:
        cursor.execute(
            "SELECT * FROM documents WHERE id = %s",
            (document_id,)
        )
        document = cursor.fetchone()

        if not document:
            raise HTTPException(
                status_code=404,
                detail="Document not found"
            )

        cursor.execute(
            "SELECT term, value FROM terms WHERE document_id = %s ORDER BY id",
            (document_id,)
        )
        terms = cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Term Grid"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="1E40AF",
            end_color="1E40AF",
            fill_type="solid"
        )
        header_alignment = Alignment(horizontal="left", vertical="center")
        alt_fill = PatternFill(
            start_color="EFF6FF",
            end_color="EFF6FF",
            fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        ws["A1"] = "Document"
        ws["B1"] = document["filename"]
        ws["A1"].font = Font(bold=True)

        ws["A2"] = "Processed"
        ws["B2"] = str(document["uploaded_at"])
        ws["A2"].font = Font(bold=True)

        ws["A3"] = "Total Terms"
        ws["B3"] = len(terms)
        ws["A3"].font = Font(bold=True)

        ws.append([])
        ws.append(["Term", "Value"])

        header_row = ws[5]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for i, term in enumerate(terms):
            row_num = i + 6
            ws.append([term["term"], term["value"]])
            for cell in ws[row_num]:
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True
                )
                if i % 2 == 1:
                    cell.fill = alt_fill

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 65
        ws.row_dimensions[5].height = 25
        for i in range(len(terms)):
            ws.row_dimensions[i + 6].height = 40

        ws.freeze_panes = "A6"

        safe_filename = document["filename"].replace(".", "_")
        export_path = f"exports/{safe_filename}_terms.xlsx"
        os.makedirs("exports", exist_ok=True)
        wb.save(export_path)

        return FileResponse(
            path=export_path,
            filename=f"{safe_filename}_terms.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise
    except Exception as e:
        
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents/compare")
def compare_documents(doc1_id: int, doc2_id: int):
    try:
        cursor.execute(
            "SELECT * FROM documents WHERE id = %s",
            (doc1_id,)
        )
        doc1 = cursor.fetchone()

        if not doc1:
            raise HTTPException(
                status_code=404,
                detail=f"Document {doc1_id} not found"
            )

        cursor.execute(
            "SELECT * FROM documents WHERE id = %s",
            (doc2_id,)
        )
        doc2 = cursor.fetchone()

        if not doc2:
            raise HTTPException(
                status_code=404,
                detail=f"Document {doc2_id} not found"
            )

        cursor.execute(
            "SELECT term, value FROM terms WHERE document_id = %s ORDER BY id",
            (doc1_id,)
        )
        terms1 = cursor.fetchall()

        cursor.execute(
            "SELECT term, value FROM terms WHERE document_id = %s ORDER BY id",
            (doc2_id,)
        )
        terms2 = cursor.fetchall()

        terms1_dict = {
            normalize_term_key(t["term"]): t["value"] for t in terms1
        }
        terms2_dict = {
            normalize_term_key(t["term"]): t["value"] for t in terms2
        }

        all_terms = sorted(set(
            list(terms1_dict.keys()) + list(terms2_dict.keys())
        ))

        comparison = []
        for term_key in all_terms:
            value1 = terms1_dict.get(term_key, "—")
            value2 = terms2_dict.get(term_key, "—")

            original_term = next(
                (t["term"] for t in terms1
                 if normalize_term_key(t["term"]) == term_key),
                next(
                    (t["term"] for t in terms2
                     if normalize_term_key(t["term"]) == term_key),
                    term_key
                )
            )

            match = "Match" if normalize(value1) == normalize(value2) \
                else "No Match"

            comparison.append({
                "term": original_term,
                "value1": value1,
                "value2": value2,
                "match": match
            })

        return {
            "doc1": doc1,
            "doc2": doc2,
            "comparison": comparison,
            "total": len(comparison),
            "matches": sum(1 for c in comparison if c["match"] == "Match"),
            "differences": sum(
                1 for c in comparison if c["match"] == "No Match"
            )
        }

    except HTTPException:
        raise
    except Exception as e:
        
        raise HTTPException(status_code=500, detail=str(e))
    

@app.post("/api/documents/cancel")
async def cancel_upload():
    cancel_processing()
    return {"message": "Cancellation requested"}